# -*- coding: utf-8 -*-
"""
出单统计日报 → 飞书多维表格同步

用法:
    python sync.py --file 出单统计.xls              # 同步
    python sync.py --file 出单统计.xls --dry-run    # 只看解析结果和差异，不写飞书
    python sync.py --dir D:\\报表 --dry-run          # 取目录里最新的 xls/xlsx
    python sync.py --file x.xls --date 2026-08-06   # 只同步指定统计周期的行

唯一键 = 统计周期 + 用户ID，重复执行同一天的文件只会更新不会重复写。

凭据从环境变量读，其次 config.json（见 config.example.json）：
    FEISHU_APP_ID / FEISHU_APP_SECRET / BITABLE_APP_TOKEN / BITABLE_TABLE_ID
"""

import argparse
import glob
import json
import os
import sys
import time
from datetime import datetime, date
from decimal import Decimal

import requests

FEISHU_BASE = "https://open.feishu.cn"
BATCH_SIZE = 500          # 飞书 batch_create/batch_update 单次上限
REQUEST_TIMEOUT = 30
MAX_RETRY = 3

# ======================== 报表列定义 ========================

# 出单统计表的 16 列，顺序与导出文件一致
COLUMNS = [
    "报表类型", "统计周期", "用户ID", "账户等级", "商户名称", "直签人",
    "代理商用户ID", "代理商名称", "站点", "接入模式",
    "交易金额（美元）", "环比增长金额（美元）", "金额环比",
    "交易笔数", "环比增长笔数", "笔数环比",
]

# 必须当字符串处理：用户ID 是 19 位，走 float 会精度丢失
ID_COLUMNS = {"用户ID", "代理商用户ID"}

NUMBER_COLUMNS = {
    "交易金额（美元）", "环比增长金额（美元）", "金额环比",
    "交易笔数", "环比增长笔数", "笔数环比",
}

# 唯一键：同一天同一个商户只应有一条
KEY_COLUMNS = ("统计周期", "用户ID")

# 判定表头行时，命中其中几个就认为这行是表头
HEADER_ANCHORS = {"报表类型", "统计周期", "用户ID", "商户名称", "站点"}
HEADER_SCAN_ROWS = 10     # 表头前面可能有标题行，往下找几行

# 飞书多维表格字段类型
FT_TEXT = 1
FT_NUMBER = 2
FT_SINGLE_SELECT = 3
FT_MULTI_SELECT = 4
FT_DATE = 5


class SyncError(RuntimeError):
    pass


# ======================== 取值清洗 ========================

def norm_header(s):
    """表头归一化：去空白换行，半角括号统一成全角，容忍导出端的小差异"""
    s = str(s if s is not None else "").strip()
    for ch in ("\n", "\r", "\t", " ", "\u3000"):
        s = s.replace(ch, "")
    return s.replace("(", "（").replace(")", "）")


def clean_uid(x):
    """清理用户ID：不经过 float，避免 19 位大整数精度丢失"""
    if x is None:
        return ""
    if isinstance(x, int) and not isinstance(x, bool):
        return str(x)
    s = str(x).strip()
    if s in ("", "nan", "None"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    if "e" in s.lower():
        # 科学计数法用 Decimal 精确还原
        try:
            s = str(int(Decimal(s)))
        except Exception:
            pass
    return s


def uid_is_lossy(v):
    """
    用户ID 是 19 位，超过 float64 的 53 位有效精度。
    单元格一旦是「数字」格式，值在读到这里之前就已经被 Excel/解析库压成 double 了，
    末几位不可恢复（1828351428978479107 → ...479104）。
    这里只能识别并告警，让导出端把该列改成文本。
    """
    if isinstance(v, float) and abs(v) >= 2 ** 53:
        return True
    if isinstance(v, int) and not isinstance(v, bool) and abs(v) >= 2 ** 53:
        return True
    return False


def to_float(val):
    if val is None or str(val).strip() in ("", "nan", "None", "-"):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace(",", "")
    pct = s.endswith("%")
    if pct:
        s = s[:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    # "-8.84%" / 100 会得到 -0.08839999999999999，收一下尾巴
    return round(v / 100, 10) if pct else v


def to_date_str(val):
    """统计周期归一化成 YYYY-MM-DD 字符串"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def date_str_to_ms(s):
    """YYYY-MM-DD → 毫秒时间戳，写飞书日期字段用"""
    try:
        return int(datetime.strptime(s, "%Y-%m-%d").timestamp() * 1000)
    except (ValueError, TypeError):
        return None


# ======================== 读 Excel ========================

def _rows_from_xls(path):
    """老 .xls（含 WPS 导出的 OLE2）用 xlrd 读。LibreOffice 打不开这类文件。"""
    try:
        import xlrd
        from xlrd import xldate
    except ImportError:
        raise SyncError("读取 .xls 需要 xlrd，请执行: pip install xlrd")

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            v = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    v = xldate.xldate_as_datetime(v, book.datemode)
                except Exception:
                    pass
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                v = None
            row.append(v)
        rows.append(row)
    return rows


def _rows_from_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        raise SyncError("读取 .xlsx 需要 openpyxl，请执行: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def read_report(path):
    """解析出单统计表，返回 [{列名: 值}]，值已按类型清洗过"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        raw = _rows_from_xls(path)
    elif ext in (".xlsx", ".xlsm"):
        raw = _rows_from_xlsx(path)
    else:
        raise SyncError(f"不支持的文件格式: {ext}（只支持 .xls / .xlsx）")

    if not raw:
        raise SyncError("文件是空的")

    # 定位表头行：报表前面可能有标题行
    header_idx = None
    for i, row in enumerate(raw[:HEADER_SCAN_ROWS]):
        names = {norm_header(c) for c in row}
        if len(names & HEADER_ANCHORS) >= 3:
            header_idx = i
            break
    if header_idx is None:
        raise SyncError(
            f"前 {HEADER_SCAN_ROWS} 行里找不到表头（需要包含 报表类型/统计周期/用户ID 等列）"
        )

    headers = [norm_header(c) for c in raw[header_idx]]
    known = {norm_header(c) for c in COLUMNS}
    unknown = [h for h in headers if h and h not in known]
    missing = [c for c in COLUMNS if norm_header(c) not in set(headers)]
    if unknown:
        print(f"  ℹ️ 报表新增了脚本不认识的列（会一并读入）: {unknown}")
    if missing:
        print(f"  ⚠️ 报表缺少这些预期列: {missing}")

    records = []
    lossy_ids = []
    for row in raw[header_idx + 1:]:
        if not any(str(v).strip() for v in row if v is not None):
            continue                                    # 跳过空行
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            if h in ID_COLUMNS:
                if uid_is_lossy(v):
                    lossy_ids.append(clean_uid(v))
                rec[h] = clean_uid(v)
            elif h == "统计周期":
                rec[h] = to_date_str(v)
            elif h in NUMBER_COLUMNS:
                rec[h] = to_float(v)
            else:
                rec[h] = "" if v is None else str(v).strip()
        # 合计行之类没有用户ID的，不入库
        if not rec.get("用户ID"):
            continue
        records.append(rec)

    if lossy_ids:
        print(f"  ❌ 有 {len(lossy_ids)} 个用户ID是以「数字」格式存的，末几位已经被精度截断，"
              f"无法还原（例: {lossy_ids[0]}）")
        print("     这会导致写进多维表格的是错误ID、且跟历史记录对不上。"
              "请把导出报表的「用户ID」列改成文本格式再跑。")
        raise SyncError("用户ID 精度已丢失，拒绝写入错误数据")

    return records


def pick_latest_file(directory):
    files = []
    for pat in ("*.xls", "*.xlsx"):
        files.extend(glob.glob(os.path.join(directory, pat)))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        raise SyncError(f"目录里没有 xls/xlsx: {directory}")
    return max(files, key=os.path.getmtime)


# ======================== 飞书 API ========================

class Feishu:
    def __init__(self, app_id, app_secret, app_token, table_id):
        self.app_id = app_id
        self.app_secret = app_secret
        self.app_token = app_token
        self.table_id = table_id
        self._token = None
        self._base = f"{FEISHU_BASE}/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}"

    # ---------- token ----------

    @property
    def token(self):
        if self._token is None:
            url = f"{FEISHU_BASE}/open-apis/auth/v3/tenant_access_token/internal"
            data = self._request("POST", url, json={
                "app_id": self.app_id, "app_secret": self.app_secret
            }, auth=False)
            self._token = data["tenant_access_token"]
        return self._token

    def _request(self, method, url, auth=True, **kw):
        headers = kw.pop("headers", {})
        if auth:
            headers["Authorization"] = f"Bearer {self.token}"
        last = None
        for attempt in range(MAX_RETRY):
            try:
                resp = requests.request(
                    method, url, headers=headers, timeout=REQUEST_TIMEOUT, **kw
                )
                data = resp.json()
            except Exception as e:
                last = e
                print(f"  ⚠️ 请求异常，重试 {attempt + 1}/{MAX_RETRY}: {e}")
                time.sleep(2 ** attempt)
                continue
            if data.get("code") == 0:
                return data.get("data", {})
            raise SyncError(f"飞书接口返回错误 code={data.get('code')} msg={data.get('msg')}")
        raise SyncError(f"请求失败（已重试 {MAX_RETRY} 次）: {last}")

    # ---------- 字段 ----------

    def get_fields(self):
        """返回 {字段名: 字段类型}。用来过滤掉表里不存在的列，并按类型转值。"""
        fields, page_token = {}, None
        while True:
            params = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self._request("GET", f"{self._base}/fields", params=params)
            for f in data.get("items") or []:
                fields[f.get("field_name")] = f.get("type")
            page_token = data.get("page_token")
            if not data.get("has_more") or not page_token:
                break
        return fields

    # ---------- 记录 ----------

    def search_by_period(self, periods, period_type):
        """
        只捞本次涉及的统计周期的记录，避免整表扫描。
        表一天天涨，全量拉迟早会慢到不可用。
        search 不可用时回退全量。
        """
        try:
            return self._search_by_period(periods, period_type)
        except SyncError as e:
            print(f"  ⚠️ 按统计周期检索失败（{e}），回退为全量读取")
            return self.list_all()

    def _search_by_period(self, periods, period_type):
        conditions = []
        for p in sorted(periods):
            if period_type == FT_DATE:
                ms = date_str_to_ms(p)
                if ms is None:
                    continue
                conditions.append({
                    "field_name": "统计周期", "operator": "is",
                    "value": ["ExactDate", str(ms)],
                })
            else:
                conditions.append({
                    "field_name": "统计周期", "operator": "is", "value": [p],
                })
        if not conditions:
            return {}

        out, page_token = {}, None
        while True:
            params = {"page_size": BATCH_SIZE}
            if page_token:
                params["page_token"] = page_token
            data = self._request(
                "POST", f"{self._base}/records/search", params=params,
                json={"filter": {"conjunction": "or", "conditions": conditions}},
            )
            self._collect(data.get("items") or [], out)
            page_token = data.get("page_token")
            if not data.get("has_more") or not page_token:
                break
            time.sleep(0.2)
        return out

    def list_all(self):
        out, page_token = {}, None
        while True:
            params = {"page_size": BATCH_SIZE}
            if page_token:
                params["page_token"] = page_token
            data = self._request("GET", f"{self._base}/records", params=params)
            self._collect(data.get("items") or [], out)
            page_token = data.get("page_token")
            if not data.get("has_more") or not page_token:
                break
            time.sleep(0.2)
        return out

    @staticmethod
    def _collect(items, out):
        for item in items:
            f = item.get("fields") or {}
            key = make_key(to_date_str(field_text(f.get("统计周期"))),
                           clean_uid(field_text(f.get("用户ID"))))
            if key:
                out[key] = item.get("record_id")

    def batch_create(self, records):
        created = 0
        for i in range(0, len(records), BATCH_SIZE):
            batch = records[i:i + BATCH_SIZE]
            self._request("POST", f"{self._base}/records/batch_create",
                          json={"records": [{"fields": r} for r in batch]})
            created += len(batch)
            time.sleep(0.3)
        return created

    def batch_update(self, updates):
        updated = 0
        for i in range(0, len(updates), BATCH_SIZE):
            batch = updates[i:i + BATCH_SIZE]
            self._request("POST", f"{self._base}/records/batch_update",
                          json={"records": [{"record_id": rid, "fields": f}
                                            for rid, f in batch]})
            updated += len(batch)
            time.sleep(0.3)
        return updated


def field_text(v):
    """多维表格读回来的值形态不一：文本字段可能是 str，也可能是 [{'text': ...}]"""
    if v is None:
        return ""
    if isinstance(v, list):
        if not v:
            return ""
        first = v[0]
        if isinstance(first, dict):
            return str(first.get("text", ""))
        return str(first)
    if isinstance(v, dict):
        return str(v.get("text", ""))
    return str(v)


def make_key(period, uid):
    if not period or not uid:
        return ""
    return f"{period}|{uid}"


# ======================== 组装写入字段 ========================

def build_fields(rec, table_fields):
    """
    按多维表格的真实字段类型转值。
    表里没有的列直接丢掉 —— 飞书遇到未知字段会整批报错。
    """
    out = {}
    for name, value in rec.items():
        ftype = table_fields.get(name)
        if ftype is None:
            continue
        if ftype == FT_NUMBER:
            v = value if isinstance(value, (int, float)) else to_float(value)
            if v is not None:
                out[name] = v
        elif ftype == FT_DATE:
            ms = date_str_to_ms(to_date_str(value))
            if ms is not None:
                out[name] = ms
        elif ftype in (FT_SINGLE_SELECT, FT_MULTI_SELECT):
            s = "" if value is None else str(value).strip()
            if s:                                        # 选项类不能写空串
                out[name] = [s] if ftype == FT_MULTI_SELECT else s
        else:
            if value is None:
                out[name] = ""
            elif isinstance(value, float) and value.is_integer():
                out[name] = str(int(value))              # 别写成 "1084.0"
            else:
                out[name] = str(value)
    return out


# ======================== 配置 ========================

def load_config(path):
    cfg = {}
    if path and os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        cfg = raw.get("feishu", raw)

    def pick(env_key, *cfg_path):
        v = os.environ.get(env_key)
        if v:
            return v.strip()
        node = cfg
        for k in cfg_path:
            if not isinstance(node, dict):
                return ""
            node = node.get(k)
        return str(node).strip() if node else ""

    return {
        "app_id": pick("FEISHU_APP_ID", "app_id"),
        "app_secret": pick("FEISHU_APP_SECRET", "app_secret"),
        "app_token": pick("BITABLE_APP_TOKEN", "bitable", "app_token"),
        "table_id": pick("BITABLE_TABLE_ID", "bitable", "table_id"),
    }


# ======================== 主流程 ========================

def main():
    ap = argparse.ArgumentParser(description="出单统计日报 → 飞书多维表格")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--file", help="报表文件路径（.xls / .xlsx）")
    src.add_argument("--dir", help="报表目录，取其中修改时间最新的文件")
    ap.add_argument("--date", help="只同步该统计周期的行，格式 YYYY-MM-DD")
    ap.add_argument("--config", default=os.path.join(os.path.dirname(__file__), "config.json"),
                    help="配置文件路径（默认同目录 config.json）")
    ap.add_argument("--dry-run", action="store_true",
                    help="只解析并打印，不写飞书。不需要凭据也能跑")
    args = ap.parse_args()

    path = args.file or pick_latest_file(args.dir)
    print(f"📖 读取报表: {path}")
    records = read_report(path)
    print(f"  解析出 {len(records)} 条")

    if args.date:
        records = [r for r in records if r.get("统计周期") == args.date]
        print(f"  按统计周期 {args.date} 过滤后 {len(records)} 条")
    if not records:
        print("⚠️ 没有可同步的数据，退出")
        return 0

    # 同一天同一商户在报表里重复出现时，后面的覆盖前面的
    deduped, dup = {}, 0
    for r in records:
        k = make_key(r.get("统计周期"), r.get("用户ID"))
        if not k:
            continue
        if k in deduped:
            dup += 1
        deduped[k] = r
    if dup:
        print(f"  ⚠️ 报表内有 {dup} 条重复的 统计周期+用户ID，保留最后一条")

    periods = sorted({r["统计周期"] for r in deduped.values() if r.get("统计周期")})
    print(f"  涉及统计周期: {', '.join(periods) or '(空)'}")

    if args.dry_run:
        return dry_run(deduped)

    cfg = load_config(args.config)
    missing = [k for k, v in cfg.items() if not v]
    if missing:
        print(f"❌ 缺少配置: {missing}")
        print("   用环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET / "
              "BITABLE_APP_TOKEN / BITABLE_TABLE_ID，或填 config.json")
        return 1

    fs = Feishu(cfg["app_id"], cfg["app_secret"], cfg["app_token"], cfg["table_id"])

    print("\n📋 读取多维表格字段...")
    table_fields = fs.get_fields()
    print(f"  表里有 {len(table_fields)} 个字段")

    for key_col in KEY_COLUMNS:
        if key_col not in table_fields:
            print(f"❌ 多维表格缺少唯一键字段「{key_col}」，无法去重，中止")
            return 1

    absent = [c for c in COLUMNS if c not in table_fields]
    if absent:
        print(f"  ⚠️ 这些报表列在多维表格里没有对应字段，本次不会写入: {absent}")

    print("\n📖 读取已有记录...")
    existing = fs.search_by_period(set(periods), table_fields.get("统计周期"))
    print(f"  命中 {len(existing)} 条已有记录")

    creates, updates = [], []
    for key, rec in deduped.items():
        fields = build_fields(rec, table_fields)
        rid = existing.get(key)
        if rid:
            updates.append((rid, fields))
        else:
            creates.append(fields)

    if creates:
        print(f"\n📝 新增 {len(creates)} 条...")
        print(f"  ✅ 成功 {fs.batch_create(creates)} 条")
    else:
        print("\n📝 无新增")

    if updates:
        print(f"📝 更新 {len(updates)} 条...")
        print(f"  ✅ 成功 {fs.batch_update(updates)} 条")
    else:
        print("📝 无更新")

    print("\n✅ 同步完成")
    return 0


def dry_run(deduped):
    print("\n🧪 dry-run：以下内容不会写入飞书\n")
    preview = list(deduped.values())[:5]
    for i, rec in enumerate(preview, 1):
        print(f"--- 第 {i} 条  key={make_key(rec.get('统计周期'), rec.get('用户ID'))}")
        for col in COLUMNS:
            if col in rec:
                print(f"    {col}: {rec[col]!r}")
        extra = [k for k in rec if k not in COLUMNS]
        for col in extra:
            print(f"    {col}: {rec[col]!r}   (报表新增列)")
        print()
    if len(deduped) > len(preview):
        print(f"... 其余 {len(deduped) - len(preview)} 条略\n")

    amount = sum(r.get("交易金额（美元）") or 0 for r in deduped.values())
    count = sum(r.get("交易笔数") or 0 for r in deduped.values())
    print(f"合计: {len(deduped)} 个商户 / 交易金额 ${amount:,.2f} / 交易笔数 {count:,.0f}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except SyncError as e:
        print(f"\n❌ {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        sys.exit(130)
